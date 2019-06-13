"""Project DataMan

This is the code handling each of the pages.
    It connects the templates to the urls and
    the information from the database."""

from django.shortcuts import render, redirect
from django.views.generic import ListView
from django.views.generic import DetailView
from django.http import HttpResponseRedirect
from . import forms
import re
from Records.models import *
from django.forms.models import model_to_dict
from django_tables2 import RequestConfig
from Records.tables import *
from Records.read_maps import *
from datetime import datetime
from django.forms import formset_factory
import openpyxl
from openpyxl.utils import get_column_letter
import json
from os.path import basename
from django.conf import settings
from django.core.management import call_command
from django.core.files.storage import default_storage
from openpyxl.worksheet.datavalidation import DataValidation
from os import remove

from apscheduler.schedulers.background import BackgroundScheduler
scheduler = BackgroundScheduler()
job = None

from boxsdk import JWTAuth
from boxsdk import Client

import logging
info_logger = logging.getLogger("info_logger")
debug_logger = logging.getLogger("debug_logger")
error_logger = logging.getLogger("error_logger")

NEW = 'NEW'
EXISTING = '(Existing)'

"""Records named 'records'
    Give options:
     -view Archives
     -add records"""
def records(request):
    return render(request, 'records.html')

def about(request):
    info_logger.info(' Someone went to the about page!')
    return render(request, 'about.html')

"""Archive home named 'archive'
    Allow type selection"""
def archive(request):
    #get the url parameter for what type of data to display, if any
    param = re.sub('/records/archive/','',request.get_full_path())
    param = re.sub('\?type=','',param)
    if param == 'experiments' :
        model = Experiment
    elif param == 'samples':
        model = Sample
    elif param == 'datasets':
        model = Dataset
    else:
        model = None
    if model is not None:
        query_results = model.objects.all()
        results = [model_to_dict(r) for r in query_results]
        col_names = [re.sub(r'\w+\.\w+\.', '', str(i)) for i in model._meta.get_fields()]
        col_names = col_names[1:]  #using the [1:] so it skips the first item which just declares it is one to many or one to one
        fields = [model._meta.get_field(i).verbose_name for i in col_names]
    else:
        results = []
        col_names = []
        fields = []
    data = {
    'type' : param,
    'show_table': param not in ['experiments','samples','datasets'],
    'query_results': results,
    'col_names': fields,
    }
    """or:
    for type in RecordTypes:
        data[type] = type.objects.all().count()
        #or whatever you want this page to know"""
    return render(request, 'archive.html', context = data)
def create_new(request):
    return render(request, 'create-new.html',)
def success(request, message = 'Successfully recorded'):
    if request.session.get('message') and request.session['message'] != None:
        message = request.session['message']
    context = {'message':message}
    request.session['message'] = None
    return render(request, 'success.html', context)

def make_backup():
    filename = 'dump-' +datetime.today().strftime('%Y-%m-%d')+'.json'
    #creates local backup
    if settings.BACKUP_LOCATION:
        filepath = settings.BACKUP_LOCATION + filename
    file = open(filepath, 'w')
    call_command('dumpdata', 'Records', stdout=file)

    #backs up to box
    if settings.BOX_CONFIG:
        print ("Trying Upload To Box")

        #https://github.com/box/box-python-sdk/blob/master/docs/usage/

        folder_id = settings.BOX_BACKUP_LOCATION
        sdk = JWTAuth.from_settings_file(settings.BOX_CONFIG)
        client = Client(sdk)

        #if it's been uploaded already,
        #the file by that name will be in the folder
        #if it hasn't, we won't have problems uploading it
        file_id = None
        items = client.folder(folder_id=folder_id).get_items()
        for i in items:
            if i.name == filename:
                file_id = i.id

        if file_id:#it's been uploaded already
            box_file = client.file(file_id).unlock()
            box_file.update_contents(filepath)
        else:
            box_file = client.folder(folder_id).upload(filepath, filename)

        box_file.lock()

        """Push the error Logs to box as well."""
        if settings.BOX_LOGS:
            ###Doesn't yet work, I think because the files are open
            #I think I can make it work by copying them.
            log_folder_id = settings.BOX_LOGS
            for log in settings.LOG_FILES:
                filename = basename(log)
                open(filename, 'wb').write(open(log, 'rb').read())

                items = client.folder(folder_id=log_folder_id).get_items()
                file_id = None
                for i in items:
                    if i.name == filename:
                        file_id = i.id

                if file_id:#it's been uploaded already
                    box_file = client.file(file_id).unlock()
                    box_file.update_contents(filename)
                else:
                    box_file = client.folder(log_folder_id).upload(filename, filename)

                box_file.lock()
            remove(filename)
    return True

def start_job():
    global job

    #Checks if the job has already been assigned so as to avoid duplicates
    if not 'make_backup_job' in str(scheduler.get_jobs()):
        job = scheduler.add_job(make_backup, 'interval', hours = 1, id = 'make_backup_job')
    if not scheduler.running: scheduler.start()
def backup(request):
    start_job()
    options = {'Restore'}

    if settings.BOX_CONFIG:
        folder_id = settings.BOX_BACKUP_LOCATION
        sdk = JWTAuth.from_settings_file(settings.BOX_CONFIG)
        client = Client(sdk)
        items = client.folder(folder_id=folder_id).get_items()
        remote_files = []
        for i in items: remote_files.append((i.id, i.name))
    else: remote_files = []

    form = forms.BackUpSelectForm(remote_files=remote_files)
    context = {
        'header': 'Backup Options',
        'options':options,
        'form':form,
    }
    if request.method == 'POST' and request.POST.get('option') == 'Restore':
        form = forms.BackUpSelectForm(request.POST)
        if form.is_valid():
            data = form.data
            source = data['source']
            try:
                if source == 'Local':
                    file = data['file']
                    restore(file)
                elif source == 'Box' and settings.BOX_CONFIG:
                    remote_file = data['remote_file']
                    box_file = client.file(remote_file)
                    new_name = 'remote-copy-'+ box_file.get().name
                    if settings.BACKUP_LOCATION:
                        new_name = settings.BACKUP_LOCATION + new_name
                    file = open(new_name, 'wb')
                    box_file.download_to(file)
                    file.close()
                    restore(new_name)
                    remove(new_name)
                else:
                    context['error'] = 'Restore from Box unavailable.'
            except:
                context['error'] = 'Error in restore. Please retry.'
        else: context['error'] = 'Please select a backup.'

    if request.method == 'GET' and request.GET.get('option') == 'backup-now':
        make_backup()
        return redirect('backup')
    return render(request, 'backup.html', context)
def restore(filename):
    call_command('loaddata', filename, app_label='Records')

#this method handles the upload page and
#directs the sheet to the read-in method
def upload(request, option = None):
    form = forms.UploadFileForm()
    upload_status = ''
    summary = ''
    upload_summary = [("Upload summary:")]
    upload_options = {}

    print("\nUpload Page")

    if request.method == 'POST' and request.POST.get('Submit') == 'Submit':

        print ("\nAttempting Upload")
        form = forms.UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['_File']
            data = form.save(commit = False)
            lead = data.lead
        #try: #Catch invalid formats, etc.
        if True: #Allows for effective debugging
            wb = openpyxl.load_workbook(file, data_only=True)
            #read_only = True sometimes causes sharing violations
            #because it doesn't close fully
            if wb['Input']['I3'].value == "Mass spec":
                read_map = read_in_map_MS
                upload_summary = read_data(request, wb, lead, read_in_map_MS, []) #to be used in template
            elif wb['Input']['I3'].value == "Instrument Type":
                upload_summary = read_Individuals(wb[read_in_map_gen['wsIndiv']], read_in_map_gen, upload_summary)
                upload_summary = read_data(request, wb, lead, read_in_map_gen, upload_summary)#"Upload Successful"
            else:
                upload_summary = ["Unknown format. Please use one of the provided templates."]
            wb.close()
            upload_options = {
                'Confirm': True,
                'Cancel': False,
            }
            request.session['upload_status'] = upload_status
        #except:
        #    upload_status = "Read in error.\nPlease use one of the provided templates."
        #    upload_summary = ["Unknown format. Please use one of the provided templates."]
        #    request.session['upload_status'] = upload_status
        print("finished Upload")
        #print(upload_status)
        #print(upload_summary)
        #"""

        #saves summary of changes for report till confirm/delete option
        i = 0
        request.session['upload_summary'] = {}
        for report in upload_summary:
            request.session['upload_summary'][i] = report
            i +=1
        request.session.modified = True

        return redirect ('upload_confirm')
        """print("read in complete")
        if 'missing_fields' in request.session:
            del request.session['missing_fields']
        if 'missing_fields_data' in request.session:
            del request.session['missing_fields_data']
        missing_fields= get_missing_fields(wb, read_map)
        missing_fields = list(dict.fromkeys(missing_fields))
        print ('missing_fields: ', missing_fields)

        request.session['missing_fields'] = missing_fields
        return redirect ('process_missing_fields')#"""

    context = {
        'form':form,
        'upload_status':upload_status,
        'summary': summary,
        'upload_options': upload_options,
    }

    #Cancel options - currently functions
    #on keep or delete
    return render(request, 'upload.html', context)
def read_data(request, wb, lead, read_map, upload_summary):
    missing_fields = []
    summary = upload_summary

    wsIn = wb[read_map['wsIn']]
    if wsIn[read_map['start_loc']].value is None:
        return ["Empty file"]

    wlrowNum = read_map['wlrowNumInit']
    if read_map['variable_colums_TF']:
        rows = wsIn[(read_map['in_section']).format(get_column_letter(wsIn.max_column), wsIn.max_row)]
    else: rows = wsIn[(read_map['in_section']).format(wsIn.max_row)]

    #The generic sheet has the option of reading
    #in one experiment per sheet with all data
    if 'exp_name' in read_map:
        exp_name = wsIn[read_map['exp_name']].value
        if exp_name:
            #this is where we start reading and checking
            lead = wsIn[read_map['lead']].value
            if lead == None:
                missing_fields.append('lead')
            if 'IRB' in read_map:
                IRB = wsIn[read_map['IRB']].value
                try: IRB = int(IRB)
                except: IRB = None
            else: IRB = None
            if 'team' in read_map:
                team = ''
                for i in read_map['team']:
                    if wsIn[i].value != None:
                        team += str(wsIn[i].value)
                        team += ', '
            else: team = None
            if 'description' in read_map:
                des = wsIn[read_map['description']].value
            else: des = None
            e_n = exp_exist_or_new(exp_name, lead, team=team, IRB = IRB, description = des)
        summary.append(e_n)

    #Get the extra fields...

    #read samples and experiment from Input sheet
    #then read datasets from "Worklist" sheet
    #separating out QC
    for i in rows:
        # i in wsIn['B34:H63']: #.format(wsIn.max_row):
        #each record a dataset associated with a sample
        if i[read_map['sample_name']].value is None:
            break
            #All done with samples

        #Otherwise read it in
        if read_map['experiment_global']:
            e_n = exp_exist_or_new(wsIn[read_map['experiment_loc']].value, lead)
        else:
            e_n = exp_exist_or_new(i[int(read_map['experiment_loc'])].value, lead)
        experiment = Experiment.objects.all().get(_experimentName = e_n[2])
        summary.append(e_n)

        e_n = sample_exists_or_new(i[read_map['sample_name']].value, experiment, i, wsIn, read_map)
        sample = Sample.objects.all().get(_sampleName = e_n[2])
        summary.append(e_n)

    wsWL = wb[read_map['wsWL']]
    wlRows = wsWL[(read_map['wlRows']).format(wsWL.max_row)]
    inRows = wsIn[(read_map['in_section_lookup']).format(wsIn.max_row)]
    for wlRow in wlRows:
        if wlRow[read_map['dataset_name']].value is None:
            break #there isn't a dataset there
        #read in datasets
        sample_type = wlRow[read_map['wl_sample_type']].value
        if sample_type.startswith('QC'):
            qc_type = sample_type[3:]
            sample_name = str(lead+" lab QC "+qc_type)
            if Sample.objects.all().filter(_sampleName = sample_name).exists():
                sample = Sample.objects.all().get(_sampleName = sample_name)
                experiment = sample.experiment()
            else:
                if not 'QC_exp' in read_map:
                    exp_name = (str(lead+" lab QC"))
                else: exp_name = str(wsWL[read_map['QC_exp']].value)
                #or wsIn if that's where QC information is defined'
                if Experiment.objects.all().filter(_experimentName = exp_name).exists():
                    experiment = Experiment.objects.all().get(_experimentName =exp_name,  _projectLead = lead)
                    summary.append(["(DEFAULT)", 'QC Experiment: ', exp_name])
                else:
                    experiment = Experiment( _experimentName = exp_name, _projectLead =  lead)
                    experiment.save()
                sample = Sample(
                    _sampleName = sample_name,
                    _storageCondition = "QC",
                    _experiment = experiment,
                    _storageLocation = "QC",
                    _organism = "QC",
                    )
                sample.save()
                summary.append(["(DEFAULT)", 'QC Sample: ', sample_name])
            sampleRow = []
        else: #Not a QC - it's a sample defined on input
            sampleNum = wlRow[read_map['wl_sample_num']].value
            sampleRow = findIn(sampleNum, inRows, read_map['lookup_column'])
            sample_name = sampleRow[read_map['lookup_sample']].value
            if Sample.objects.all().filter(_sampleName = sample_name).exists():
                sample = Sample.objects.all().get(_sampleName = sample_name)
            else: raise ValueError("No known sample")
            experiment = sample.experiment()
        #At this point both the sample and the experiment have been defined
        #"""

        e_n = dataset_exists_or_new(wlRow[read_map['dataset_name']].value, experiment, sample, sampleRow, wb, wsIn, wlRow, read_map)
        if sample_type == 'QC': e_n[1] = "QC Dataset: "
        summary.append(e_n)

    wlRow = None
    wsIn = None
    return summary
def findIn(val, rows, lookup_column):
    for line in rows:
        if line[lookup_column].value == val:
            return line
    return []
def upload_confirm(request, option = None):
    try: upload_status = request.session['upload_status']
    except: upload_status = []
    try: summary = request.session.get('upload_summary')
    except: summary = []
    upload_summary = []
    upload_by_types = {}

    for i in summary:
        upload_summary.append(summary[i])
        if (i) != '':
            record_type = summary[i][1]
            if record_type not in upload_by_types:
                upload_by_types[record_type] = []
            upload_by_types[record_type].append(summary[i])

    if len(upload_summary) >1: summary = upload_summary[1:]
    upload_status = upload_summary[0]

    exp_table_exists = False
    sample_table_exists = False
    dataset_table_exists = False

    for i in upload_by_types:
        e_n_list = upload_by_types[i]
        this_all = []
        for e in e_n_list:
            this_all.append(e[2])
        these = []
        [these.append(x) for x in this_all if x not in these]

        if "QC" in i:
            print (i, "isn't yet a table")

        elif "Experiment" in i:
            experiment_set = Experiment.objects.all().filter(_experimentName__in=these)
            exp_table = ExperimentTable(experiment_set.order_by('-_experimentName'))
            exp_table_exists = True
        elif "Sample" in i:
            #separate QC so that doesn't overwrite set
            queryset = Sample.objects.filter(_sampleName__in = these)
            sample_table = SampleTable(queryset.order_by('-pk'))
            sample_table_exists = True
        elif "Dataset" in i:
            queryset = Dataset.objects.filter(_datasetName__in = these)
            dataset_table = DatasetTable(queryset.order_by('-pk'))
            dataset_table_exists = True

        else: print ("\n\nUnknown Type: ", i)

    tables=[]
    if exp_table_exists: tables.append(exp_table)
    if sample_table_exists: tables.append(sample_table)
    if dataset_table_exists: tables.append(dataset_table)

    upload_options = {'Confirm', 'Cancel'}
    context = {}

    if request.GET.get('option') == "Confirm":
        print ("confirm")
        context['upload_status'] = 'Saved'
        context['summary'] = ''
        try: del request.session['upload_summary']
        except:
            request.session['message'] = 'Unable to verify save.'
            return redirect('success')
        request.session['upload_status'] = 'Saved.'
        return redirect('upload')
    elif request.GET.get('option') == 'Cancel':
        context['upload_status'] = 'Cancelling...'
        print ("cancel")
        try:
            #get rid of read in data
            upload_summary = request.session.get('upload_summary')

            for i in upload_summary:
                if int(i) !=0: #
                    v = upload_summary[i]
                    if v[0] == NEW:
                        #was new with this read in
                        #needs to be deleted
                        if 'Experiment' in v[1]:
                            #it's an experiment
                            experiment = Experiment.objects.all().get(_experimentName = v[2])
                            experiment.delete()
                        elif 'Dataset' in v[1]:
                            #may have been deleted with the experiment
                            if Dataset.objects.all().filter(_datasetName = v[2]).exists():
                                dataset = Dataset.objects.all().get(_datasetName = v[2])
                                dataset.delete()
                        elif 'Sample' in v[1]:
                            #it's a sample -- it may have been deleted with the parent
                            #experiment, so we check if it exists
                            if Sample.objects.all().filter(_sampleName = v[2]).exists():
                                s = Sample.objects.all().get(_sampleName = v[2])
                                s.delete()
                        else: print (v, "Delete Unsuccessful")
            request.session['upload_status'] = 'Upload Cancelled'
            del request.session['upload_summary']
        except Exception as e:
            error_logger.error(repr(e))
            request.session['message'] = 'No upload found.'
        return redirect('upload')

    context = {
        'tables':tables,
        'summary': summary,
        'upload_options': upload_options,
    }
    if upload_status != "Upload summary:":
        context['upload_status'] = upload_status

    #Cancel options - currently functions
    #on keep or delete
    return render(request, 'upload.html', context)
def read_Individuals(wsInd, read_map, upload_summary):
    exp_n = wsInd[read_map['indivExp']].value
    print (exp_n)

    if Experiment.objects.all().filter(_experimentName = exp_n).exists():
        exp = Experiment.objects.all().get(_experimentName = exp_n)
    else: return upload_summary

    rows =  wsInd[(read_map['indivRows']).format(get_column_letter(wsInd.max_column), wsInd.max_row)]

    indivs = Individual.objects.all()
    for i in rows:
        indivID = i[read_map['indivID']].value
        if not indivID: break
        if Individual.objects.all().filter(_individualIdentifier = indivID).exists():
            upload_summary.append([EXISTING, 'Individual:', indivID])
            break

        new_Ind = Individual(
            _individualIdentifier = indivID,
            _experiment = exp,
            _gender = i[read_map['gender']].value,
            _age = i[read_map['age']].value,
            _healthStatus = i[read_map['health_status']].value,
            _comments = i[read_map['indivComments']].value
        )

        upload_summary.append([NEW, 'Individual:', indivID])

        new_Ind.save()

    return upload_summary

###NOT IMPLEMENTED YET 3-19-19###
def get_missing_fields(wb, read_map):
    missing_fields = []
    if 'missing_fields' in read_map:
        missing_fields = read_map['missing_fields']
    #the rest of our checking
    #Required Fields:
    #Exp. Name, Project Lead
    #    Design needs a name (exp does not need design)
    #
    #Individual ID, gender, age, health Status & extra_fields
    #
    #samples name, exp, storage_condition, location, organism
    #
    #dataset name, sample, instrument, file name, extension, path

    wsIn = wb[read_map['wsIn']]
    wsWL = wb[read_map['wsWL']]

    #bad logic
    if read_map['experiment_global']:
        name = wsIn[read_map['experiment_loc']].value
        if name == None:
            missing_fields.append('experiment_name')
        elif not experiments.filter(_experimentName = name).exists():
            #check for other experiment data
            if 'lead' in read_map:
                if wsIn[read_map['lead']].value == None:
                    missing_fields.append('lead')
            elif not 'lead' in missing_fields: missing_fields.append('lead')
        elif 'lead' in missing_fields:
            missing_fields.remove('lead')
    else:
        missing_fields.append('lead')
        missing_fields.append('IRB')
        missing_fields.append('description')

    if 'wsIndividual'in read_map:
        #check individuals
        if read_map['wsIndividual'] in wb:
            wsIndividual = wb[read_map['wsIndividual']]

    if read_map['variable_colums_TF']:
        rows = wsIn[(read_map['in_section']).format(get_column_letter(wsIn.max_column), wsIn.max_row)]
    else: rows = wsIn[(read_map['in_section']).format(wsIn.max_row)]

    #Do I now check each sample, the first one, or what?
    #Maybe the e_n functions report missing data
    #and here we just catch overall missing categories of data?

    #Now I'm wondering if I need to ditch this and make it along with the read in?

    return missing_fields
def process_missing_fields(request):
    #if not 'missing_fields' in request.session:
    #    return redirect('upload')

    missing_fields = request.session['missing_fields']
    #"""
    #testing not verified
    form = forms.ListFieldsForm(extraFields = missing_fields)
    if request.method == 'POST':
        print ("POST")
        #Checked and sent to form but not yet recorded
        form = forms.ListFieldsForm(request.POST)
        request.session['missing_field_data'] = form.data
        print (form.data)

        upload_summary = request.session['upload_summary'] #to be used in template
        del request.session['missing_fields']
        del request.session['missing_field_data']
        return redirect ('upload_confirm')

    context = {'form':form, 'header':'Please address missing fields'}

    return render(request, 'title-form.html', context)

#overload for additional information
def exp_exist_or_new(name, lead, team = None, IRB = None, description = None, ):
    experiments = Experiment.objects.all()
    if experiments.filter(_experimentName = name).exists():
        return [EXISTING, 'Experiment: ', name]
    #the experiment needs to be created
    newExp = Experiment(
        _experimentName = name,
        _projectLead =  lead,
    )
    if team:
        newExp.setTeamMembers(team)
    if IRB:
        try: newExp.setIRB(IRB)
        except ValueError: IRB = None
    if description: newExp.setComments(description)
    newExp.save()
    return [NEW, 'Experiment: ', name]# newExp.experimentID()]
def sample_exists_or_new(name, experiment, row, wsIn, read_map, extra_fields = None):
    samples = Sample.objects.all()
    if samples.filter(_sampleName = name).exists():
        return [EXISTING, 'Sample: ', name]

    comment = ""
    for c in read_map['comments_row']:
        c_val = str(row[read_map['comments_row'][c]].value)
        if c_val != None and c_val !="None":
            comment += c
            comment += c_val +'\n'
    for c in read_map['comments_gen']:
        c_val = str(wsIn[read_map['comments_gen'][c]].value)
        if c_val != None and c_val !="None":
            comment += c
            comment += c_val+'\n'

    #read in the extra fields
    if extra_fields:
        extraFieldData = {}
        i = read_map['extra_f_s_start']
        for f in extra_fields:
            extraFieldData[f] = str(row[i].value)
            i += 1

    if read_map['date_global']: date = wsIn[read_map['date_created']].value
    else: date = row[read_map['date_created']].value

    try:
        date = datetime.strptime(date, '%m-%d-%Y').strftime('%Y-%m-%d')
    except:
        date = ''

    #Check the individual
    if row[read_map['individualID']] != None:
        indivID = row[read_map['individualID']].value
        if Individual.objects.all().filter(_individualIdentifier = indivID).exists():
            ind = Individual.objects.all().get(_individualIdentifier = indivID)
        else:
            ind = Individual(_individualIdentifier = indivID)
            ind.save()
    else: ind = None

    def protocol_exists_or_new(ptName):
        if Protocol.objects.all().filter(_name = ptName).exists():
            return [EXISTING,'Protocol: ', Protocol.objects.all().get(_name = ptName)]
        prot = Protocol(_name = ptName)
        prot.save()
        return [NEW,'Protocol: ',prot]

    if read_map['storage_condition']: condition = str(row[read_map['storage_condition']].value)
    else: condition = "Unspecified"

    newSample = Sample(
        _sampleName = name,
        _storageCondition = condition,
        _experiment = experiment,
        _storageLocation = str(row[read_map['storage_location']].value),
        _organism = wsIn[(read_map['organism'])].value,
        _comments = comment,
    )
    if date !='': newSample.setDateCreated(date)
    if extra_fields: newSample.setExtraFields(extraFieldData)
    newSample.save()
    if ind: newSample.setIndividual([ind])
    newSample.save()

    if read_map['sheetType'] == 'General':
        protocolNames = str(row[read_map['protocol']].value)
        protocolNames = protocolNames.split(',')

        for i in protocolNames:
            p = protocol_exists_or_new(i.strip())[2]
            newSample._treatmentProtocol.add(p)

    return [NEW, 'Sample: ', name]
def dataset_exists_or_new(name, experiment, sample, row, wb, wsIn, wlRow, read_map):
    datasets = Dataset.objects.all()

    if datasets.filter(_datasetName = name).exists():
        return [EXISTING, 'Dataset: ', name]

    def inst_exists_or_new(insName):
        if Instrument.objects.all().filter(_name = insName).exists():
            return [EXISTING,'Instrument: ', Instrument.objects.all().get(_name = insName)]
        ins = Instrument(_name = insName)
        ins.save()
        return [NEW, 'Instrument: ', ins]
    def setting_exists_or_new(methodName):
        if InstrumentSetting.objects.all().filter(_name = methodName).exists():
            return [EXISTING,'Instrument Setting: ',InstrumentSetting.objects.all().get(_name = methodName)]
        ins = InstrumentSetting(_name = methodName)
        #If we wanted to try reverse-engineering the lookup function and get the rest of the setting information
        #it'd look a bit like this.
        #settings_column = wb[read_map['settings_sheet']][read_map['settings_keyword_column']]

        filename = wlRow[read_map['settings_file']].value
        try: ins.file = open(filename)
        except:
            if ins.comments == None:
                ins.comments = filename
            else:
                ins.comments = (str(ins.comments) + finename)
        ins.save()
        return [NEW, 'Instrument Setting: ', ins]

    summary = []
    ins_code = wsIn[read_map['inst_code']].value
    if ins_code == None: ins_code = ''
    else: ins_code = ' '+ins_code
    print("\n\n\n", ins_code)
    e_n = inst_exists_or_new(str(wsIn[read_map['instrument_type_loc']].value)+ins_code)
    initInstrument = e_n[2]
    print (initInstrument)
    if (e_n[0] == NEW): summary.append(e_n)
    if row != []:
        e_n = setting_exists_or_new(row[read_map['setting_loc']].value)
        setting = e_n[2]
    else: setting = None #wlRow[read_map['settings_file']]

    dataType = wsIn[read_map['data_type_loc']].value
    try: date = datetime.strptime(str(wsIn[read_map['date_loc']].value), '%m-%d-%Y').strftime('%Y-%m-%d')
    except: date = False

    wsWL = wb[read_map['wsWL']]
    if read_map['file_extension_from_excel']: extension = str(wlRow[read_map['file_extension']].value)
    else: extension = str(read_map['file_extension'])

    newDataset = Dataset(
        _datasetName = name,
        _experiment = experiment,
        _instrument = initInstrument,
        _fileName = str(wlRow[read_map['file_name']].value)+extension,
        _fileLocation = str(wlRow[read_map['file_location']].value),
        _fileLocationRemote = str(wlRow[read_map['file_location_remote']].value),
        _instrumentSetting = setting,
        _type = dataType,
        #Commas after each value
        #acquisition dates
        #_status
        #_size
        #_fileHash
    )
    if date: newDataset.setDateCreated(date)
    newDataset.save()
    newDataset._sample.add(sample)

    return [NEW, 'Dataset: ', name]

"""Page to add a sample"""
def add_sample(request, experiment = None):
    if experiment == None:
        form = forms.SelectExperiment()# GetExperForm()
        context = {
            'form':form,
            'header':'Add Sample',
        }
        if request.method == 'POST':
            form = forms.SelectExperiment(request.POST)
            if form.is_valid():
                f = form.save(commit = False)
                key = f.experiment()._experimentID
            return redirect('add-samples', key)
        extra = []
        return render(request, 'title-form.html', context)
    else:
        try:
            experiment_set = Experiment.objects.all()
            exp = experiment_set.get(pk = experiment)
            extra = exp.experimentalDesign().extra_fields_samples()
        except: extra = []
    #check anything you want checked
    form =forms.AddSampleForm(extraFields = extra)
    if request.method == 'POST':
        form =forms.AddSampleForm(request.POST, extraFields = extra)
        if form.is_valid():
            new_Sample = form.save(commit = False)
            #parses to JSON
            extraFieldData = {}
            for f in extra:
                extraFieldData[f] = form.data[f]
            new_Sample.setExtraFields(extraFieldData)
            new_Sample.setExperiment(exp)
            new_Sample.save()
            return redirect('samples')
    buttons = {
        'New Protocol': 'add-protocol',
    }
    context = {
        'form':form,
        'header':'Add Sample',
        'buttons':buttons
    }

    return render(request, 'add-record.html', context)
def add_dataset(request):
    form =forms.AddDatasetForm()
    if request.method == 'POST':
        form =forms.AddDatasetForm(request.POST)
        if form.is_valid():
            new_Dataset = form.save()
            exp = new_Dataset.sample()[0].experiment()
            new_Dataset.setExperiment(exp)
            new_Dataset.save()
            return redirect('datasets')
    buttons = {
        'New Instrument': 'add-instrument',
        'New Instrument Setting': 'add-instrument-setting'
    }
    context = {
        'form':form,
        'header':'Add Dataset',
        'buttons':buttons
    }

    return render(request, 'add-record.html', context)
def add_experiment(request):
    form = forms.AddExperimentForm()
    desForm = forms.ExperimentalDesignForm()
    if request.method == 'POST':
        form =forms.AddExperimentForm(request.POST)
        desForm = forms.ExperimentalDesignForm(request.POST, request.FILES)
        if form.is_valid() and desForm.is_valid():
            new_Experiment = form.save(commit = False)
            newDes = desForm.save()
            new_Experiment.setExperimentalDesign(newDes)
            new_Experiment.save()
            return redirect('experiments')
    buttons = {
        #'New Experimental Design': 'add-experimental-design',
    }
    context = {
        'header':'Add Experiment',
        'form':form,
        'secFormHeading':'Experimental Design',
        'secForm':desForm,
        'buttons':buttons
    }

    return render(request, 'add-record.html', context)
def add_individual(request, experiment = None):
    if experiment == None:
        form = forms.SelectExperiment()# GetExperForm()
        context = {
            'form':form,
            'header':'Add Individual',
        }
        if request.method == 'POST':
            form = forms.SelectExperiment(request.POST)
            if form.is_valid():
                individual = form.save(commit = False)
                key = individual.experiment()._experimentID
            return redirect('add-individual', key)
        extra = []
        return render(request, 'title-form.html', context)
    else:
        try:
            experiment_set = Experiment.objects.all()
            exp = experiment_set.get(pk = experiment)
            extra = exp.experimentalDesign().extra_fields()
        except: extra = []

    form = forms.AddIndividualForm(extraFields = extra)
    if request.method == 'POST':
        form = forms.AddIndividualForm(request.POST, extraFields = extra)
        if form.is_valid():
            new_Individual = form.save(commit = False)
            #parses to JSON
            extraFieldData = {}
            for f in extra:
                extraFieldData[f] = form.data[f]
            new_Individual.setExtraFields(extraFieldData)
            new_Individual.setExperiment(exp)
            new_Individual.save()
            return redirect('individuals')
    context = {
        'form':form,
        'header':'Add Individual',
    }

    return render(request, 'add-record.html', context)
def add_instrument(request):
    form = forms.InstrumentForm()
    if request.method == 'POST':
        form = forms.InstrumentForm(request.POST, request.FILES)
        if form.is_valid():
            new_inst = form.save()
            return redirect('add-dataset')
    context = {
        'form':form,
        'header':'Add Instrument'
    }
    return render(request, 'add-record.html', context)
def add_instrument_setting(request):
    form = forms.InstrumentSettingForm()
    if request.method == 'POST':
        form = forms.InstrumentSettingForm(request.POST, request.FILES)
        if form.is_valid():
            new_inst = form.save()
            return redirect('add-dataset')
    context = {
        'form':form,
        'header':'Add Instrument Setting'
    }
    return render(request, 'add-record.html', context)

def add_protocol(request):
    #form = forms.ProtocolForm()
    secForm = forms.FileForm()
    if request.method == 'POST':
        #form = forms.ProtocolForm(request.POST, request.FILES)
        secForm = forms.FileForm(request.POST, request.FILES)
        if secForm.is_valid(): # form.is_valid() and
            #new_inst = form.save()

            #This is getting the files, but tends to read
            #it as a bytes stream rather than multiple files
            #only last is being saved
            print("\n\n\n\nFiles")
            print (request.FILES)
            for i in request.FILES:
                print (request.FILES[i])
            #f = secForm.save()
            return redirect('add-sample')
    context = {
        #'form':form,
        'header':'Add Protocol',
        'secForm':secForm,
    }
    return render(request, 'add-record.html', context)
def add_file_status(request):
    form = forms.fileStatusOptionForm()
    if request.method == 'POST':
        form = forms.fileStatusOptionForm(request.POST, request.FILES)
        if form.is_valid():
            new_inst = form.save()
            return redirect('add-dataset')
    context = {
        'form':form,
        'header':'Add File Status'
    }
    return render(request, 'add-record.html', context)
def add_experimental_design(request):
    form = forms.ExperimentalDesignForm()
    if request.method == 'POST':
        form = forms.ExperimentalDesignForm(request.POST, request.FILES)
        if form.is_valid():
            newDes = form.save()
            return redirect('add-experiment')
    context = {
        'form':form,
        'header':'Add Experimental Design'
    }
    return render(request, 'add-record.html', context)


def edit_dataset(request, pk):
    dataset = Dataset.objects.get(pk=pk or None)
    form = forms.AddDatasetForm(instance = dataset)
    if request.method == 'POST':
        form =forms.AddDatasetForm(request.POST, instance = dataset)
        if form.is_valid():
            dataset = form.save(commit = False)
            dataset._experiment = dataset.sample()[0].experiment()
            dataset.save()
            return redirect('success')

    buttons = {
        'New Instrument':'add-instrument',
        'New Instrument Setting':'add-instrument-setting'
    }
    context = {
        'form':form,
        'header':'Edit Dataset',
        'dataset':dataset,
        'buttons':buttons
    }

    return render(request, 'add-record.html', context)
def edit_sample(request, pk):
    sample = Sample.objects.get(pk=pk)
    form = forms.AddSampleForm(instance = sample)
    if request.method == 'POST':
        form =forms.AddSampleForm(request.POST, instance = sample)
        if form.is_valid():
            sample = form.save()
            #updates dataset key to experiment as well
            try:
                if sample.dataset:
                    sample.dataset._experiment = sample.experiment()
                    sample.dataset.save()
            except:
                DoesNotExist = null #Does nothing but make it not crash
                #something needs to be here, doesn't matter what
            finally:
                return redirect('success')

    buttons = {
        'New Protocol':'add-protocol',
    }
    context = {
        'form':form,
        'header':'Edit Sample',
        'sample':sample,
        'buttons':buttons
    }

    return render(request, 'add-record.html', context)
def edit_experiment(request, pk):
    experiment = Experiment.objects.get(pk=pk)
    form = forms.AddExperimentForm(instance = experiment)
    if request.method == 'POST':
        form =forms.AddExperimentForm(request.POST, instance = experiment)
        if form.is_valid():
            experiment = form.save()
            return redirect('success')
    buttons = {
        'New Experimental Design':'add-experimental-design',
    }
    context = {
        'form':form,
        'header':'Edit Experiment',
        'experiment':experiment,
        'buttons':buttons
    }

    return render(request, 'add-record.html', context)

"""(type)Views generate the list view pages
i.e., the list of samples that's now a table"""
class SampleView(ListView):
    model = Sample

    def get_exp(self, context):
        form = forms.SelectExperiment()
        if self.request.method == 'GET':
            form =forms.SelectExperiment(self.request.GET)
            if form.is_valid():
                i = form.save(commit = False)
                context['experiment'] = i.experiment()._experimentID
        context['filter_by_exp'] = form
        return context

    def get_queryset(self, experiment = None):
        if experiment == None:
            queryset = Sample.objects.all()
        else:
            queryset = Sample.objects.filter(_experiment = experiment)
        return queryset

    #DataMan\Records\templates\Records\samples_list.html
    template_name = 'samples_list.html'
    context_object_name = 'sample'
    def get_context_data(self, **kwargs):
        context = super(SampleView, self).get_context_data(**kwargs)
        context = self.get_exp(context, **kwargs)
        if 'experiment' in context:
            queryset = self.get_queryset(context['experiment'])
        else: queryset = self.get_queryset()
        table = SampleTable(queryset.order_by('-pk'))
        RequestConfig(self.request, paginate={'per_page': 25}).configure(table)
        context['table'] = table
        context['Title'] = 'Sample'
        return context
class DatasetView(ListView):
    model = Dataset

    def get_exp(self, context):
        form = forms.SelectExperiment()
        if self.request.method == 'GET':
            form =forms.SelectExperiment(self.request.GET)
            if form.is_valid():
                i = form.save(commit = False)
                context['experiment'] = i.experiment()._experimentID
        context['filter_by_exp'] = form
        return context

    def get_queryset(self, experiment = None):
        if experiment == None:
            queryset = Dataset.objects.all()
        else:
            queryset = Dataset.objects.filter(_experiment = experiment)
        return queryset

    template_name = 'dataset_list.html'
    context_object_name = 'dataset'
    def get_context_data(self, **kwargs):
        context = super(DatasetView, self).get_context_data(**kwargs)
        context = self.get_exp(context, **kwargs)
        if 'experiment' in context:
            queryset = self.get_queryset(context['experiment'])
        else: queryset = self.get_queryset()
        table = DatasetTable(queryset.order_by('-pk'))
        RequestConfig(self.request, paginate={'per_page': 25}).configure(table)
        context['table'] = table
        context['Title'] = 'Dataset'
        return context
class ExperimentView(ListView):
    model = Experiment
    queryset = Experiment.objects.all()
    template_name = 'experiment_list.html'
    context_object_name = 'experiment'
    def get_context_data(self, **kwargs):
        context = super(ExperimentView, self).get_context_data(**kwargs)
        table = ExperimentTable(Experiment.objects.all().order_by('-_experimentName'))
        RequestConfig(self.request, paginate={'per_page': 25}).configure(table)
        context['table'] = table
        context['Title'] = 'Experiment'
        return context
class IndividualView(ListView):
    model = Individual
    queryset = Individual.objects.all()
    template_name = 'individual_list.html'
    paginate_by = 25
    def get_context_data(self, **kwargs):
        context = super(IndividualView, self).get_context_data(**kwargs)
        table = IndividualTable(Individual.objects.all().order_by('_individualIdentifier'))
        RequestConfig(self.request, paginate={'per_page': 25}).configure(table)
        context['table'] = table
        context['Title'] = 'Individuals'
        return context
class InstrumentView(ListView):
    model = Instrument
    template_name = 'instrument_list.html'
    paginate_by = 25
    def get_context_data(self, **kwargs):
        context = super(InstrumentView, self).get_context_data(**kwargs)
        RequestConfig(self.request, paginate={'per_page': 25})
        context['Title'] = 'Instruments'
        return context
class SettingsView(ListView):
    model = InstrumentSetting
    template_name = 'instrumentsetting_list.html'
    paginate_by = 25
    def get_context_data(self, **kwargs):
        context = super(SettingsView, self).get_context_data(**kwargs)
        RequestConfig(self.request, paginate={'per_page': 25})
        context['Title'] = 'Instrument Settings'
        return context
class ProtocolView(ListView):
    model = Protocol
    template_name = 'protocol_list.html'
    paginate_by = 25
    def get_context_data(self, **kwargs):
        context = super(ProtocolView, self).get_context_data(**kwargs)
        RequestConfig(self.request, paginate={'per_page': 25})
        context['Title'] = 'Sample Treatment Protocols'
        return context
class ExpDesignView(ListView):
    model = ExperimentalDesign
    template_name = 'instrument_list.html'
    paginate_by = 25
    def get_context_data(self, **kwargs):
        context = super(ExpDesignView, self).get_context_data(**kwargs)
        RequestConfig(self.request, paginate={'per_page': 25})
        context['Title'] = 'Experimental Designs'
        return context

"""named (type)-detail, these are the detail view
pages generated for the specific record
Ex: sample 1234, dataset 3"""
class SampleDetailView(DetailView):
    model = Sample
    template = 'sample_detail.html'

    def get_context_data(self, **kwargs):
        context = super(SampleDetailView, self).get_context_data(**kwargs)
        designSet = context['sample'].treatmentProtocol().all()
        for design in designSet:
            print(design)
            context['protocol'] = design
            context['protocol.description'] = design.description()
            if design.file():
                context['protocol_filename'] = basename(design.file().path)
                context['protocol_download'] =design.file().url
        return context
class DatasetDetailView(DetailView):
    instrumentSetting = "new"
    model = Dataset
    template = 'dataset_detail.html'

    def get_context_data(self, **kwargs):
        context = super(DatasetDetailView, self).get_context_data(**kwargs)
        instrument = context['dataset'].instrument()
        context['instrument'] = instrument

        context['instrument.description'] = instrument.description()
        if instrument.file():
            context['instrument_filename'] = basename(instrument.file().path)
            context['instrument_download'] =instrument.file().url

        setting = context['dataset'].instrumentSetting()
        context['setting'] = setting

        if setting != None:
            context['setting.description'] = setting.description()
            if setting.file():
                context['setting_filename'] = basename(setting.file().path)
                context['setting_download'] = setting.file().url
        return context
class ExperimentDetailView(DetailView):
    model = Experiment
    template = 'experiment_detail.html'

    def get_context_data(self, **kwargs):
        context = super(ExperimentDetailView, self).get_context_data(**kwargs)
        design = context['experiment'].experimentalDesign()
        context['design'] = design
        if design != None:
            context['design.description'] = design.description()
            if design.file():
                context['design_filename'] = basename(design.file().path)
                context['design_download'] = design.file().url
        return context
class IndividualDetailView(DetailView):
    model = Individual
    template = 'individual_detail.html'
